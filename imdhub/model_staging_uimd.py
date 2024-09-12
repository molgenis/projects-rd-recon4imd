"""Generate staging table schema
FILE: model_staging_uimd.py
AUTHOR: David Ruvolo
CREATED: 2024-08-31
MODIFIED: 2024-08-31
PURPOSE: create schema for staging table
STATUS: in.progress
PACKAGES: **see below**
COMMENTS: NA
"""

from openpyxl import load_workbook
from datatable import dt, f
import re


def sanitise_comments(value: str):
    """Clean comments"""
    return re.sub(r'(\n{1,})', '', value)


def set_column_type(value: str):
    """Set molgenis column type based on keywords"""
    if ('date' in value):
        return 'date'
    elif (value in ['age', 'head', 'height', 'weight']):
        return 'number'
    else:
        return 'string'


# open workbook and extract row 2 in first worksheet
wb = load_workbook('data/Batch_upload_TEMPLATE_Visits.xlsx')
sheet_name = wb.sheetnames[0]
data = wb[sheet_name]
target_row: tuple = data['A2:VQ2']

# ///////////////////////////////////////

# set values and comments as molgenis schema attributes

tableName = 'Batch visits'

schema: list = [{
    'tableName': tableName,
    'description': 'Staging table for Batch visits upload template'
}]

for cell in target_row[0]:
    schema_entry = {'tableName': tableName, 'columnName': cell.value}
    if (cell.comment is not None):
        schema_entry['description'] = sanitise_comments(cell.comment.text)
    schema.append(schema_entry)


# ///////////////////////////////////////


# convert to dt and add attributes
mg_schema = dt.Frame(schema)

# guess column type based on keywords (apply very basic typing)
mg_schema['columnType'] = dt.Frame([
    set_column_type(value) if value is not None else None
    for value in mg_schema['columnName'].to_list()[0]
])

# set key based on column name
mg_schema['key'] = dt.Frame([
    1 if value == 'uimd_id' else None
    for value in mg_schema['columnName'].to_list()[0]
])


mg_schema.to_csv('data/molgenis.csv')
